import streamlit as st
import requests
import jwt
from datetime import datetime
import plotly.express as px
from database import oracle, chatwoot
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
import unicodedata
import io
import xlsxwriter
import extra_streamlit_components as stx
import os
import bcrypt
import pytz
import plotly.graph_objects as go

EMAILS_PRPOSTAS = [
    "pablo.ti@caiuas.com.br",
    "cristiane.aguilar@caiuas.com.br"
]


def render():
    initial_date = st.sidebar.date_input("Data Inicial", datetime.now().date(), key="prop_data_inicial")
    final_date = st.sidebar.date_input("Data Final", datetime.now().date(), key="prop_data_final")
    
    conn, cur = oracle()
    query = f"""
        SELECT 
            to_char(vp.COD_PROPOSTA) COD_PROPOSTA,
            vp.EMISSAO data_proposta, 
            eu.NOME_COMPLETO nome_vendedor, 
            pm.DESCRICAO_MODELO modelo, 
            COALESCE(ce_veic.DESCRICAO, ce_ped.DESCRICAO, ce_fic.DESCRICAO) AS cor, 
            v.CHASSI_COMPLETO, 
            c.NOME nome_cliente,
            CASE 
                WHEN c.NUMERO_MSG_TXT_INST IS NOT NULL THEN c.PREFIXO_MSG_TXT_INST || c.NUMERO_MSG_TXT_INST
                WHEN c.TELEFONE_CEL IS NOT NULL THEN c.PREFIXO_CEL || c.TELEFONE_CEL
                WHEN c.TELEFONE_res IS NOT NULL THEN c.PREFIXO_RES || c.TELEFONE_res
                WHEN c.TELEFONE_COM IS NOT NULL THEN c.PREFIXO_COM || c.TELEFONE_COM
            END AS telefone,
            c.EMAIL_NFE email,
            CASE 
                WHEN v.novo_usado = 'U' THEN 'Usado'
                ELSE 'Novo'
            END novo_usado,
            CASE
                WHEN c.COD_CID_COM IS NOT NULL THEN cid_com.DESCRICAO 
                WHEN c.COD_CID_RES IS NOT NULL THEN cid_res.DESCRICAO 
                ELSE cid_cob.DESCRICAO 
            END cidade,
            'Aguardando Faturamento' Status,
            concat(cev.cod_empresa, cev.COD_EVENTO) evento,
            ca.ANDAMENTO andamento,
            CASE 
                WHEN cev.COD_PROPOSTA IS NOT NULL THEN 'Super quente'
                WHEN cev.TERMOMETRO = 1 THEN 'Frio'
                WHEN cev.TERMOMETRO = 2 THEN 'Morno'
                WHEN cev.TERMOMETRO = 3 THEN 'Quente'
                ELSE 'Não classificado'
            END AS termometro
        FROM VEICULOS_PROPOSTAS vp 
        LEFT JOIN produtos pr ON 1=1
            AND pr.COD_PRODUTO = vp.COD_PRODUTO  
        LEFT JOIN produtos_modelos pm ON 1=1
            AND pm.COD_PRODUTO = vp.COD_PRODUTO 
            AND pm.COD_MODELO = vp.COD_MODELO 
        LEFT JOIN VEICULOS v ON 1=1
            AND vp.CHASSI_RESUMIDO = v.CHASSI_RESUMIDO
            AND vp.COD_PRODUTO = v.COD_PRODUTO      
            AND vp.COD_MODELO = v.COD_MODELO        
            AND vp.cod_empresa = v.cod_empresa      
            AND v.status = 'E' 
        LEFT JOIN VEICULOS_PEDIDOS vped ON 1=1
            AND vp.COD_PEDIDO = vped.COD_PEDIDO
        LEFT JOIN PROP_FICTICIA_DADOS pfd ON 1=1
            AND vp.COD_FICTICIO = pfd.COD_FICTICIO
        LEFT JOIN clientes c ON c.COD_CLIENTE = vp.COD_CLIENTE 
        LEFT JOIN patio p ON 1=1
            AND p.COD_PATIO = v.COD_PATIO
        LEFT JOIN EMPRESAS_USUARIOS eu ON 1=1
            AND eu.NOME = vp.VENDEDOR 
        LEFT JOIN empresas_usuarios eu2 ON 1=1
            AND eu2.nome = vp.QUEM_APROVOU 
        LEFT JOIN empresas e ON 1=1
            AND e.cod_empresa = v.COD_EMPRESA 
        LEFT JOIN cidades cid_res ON 1=1
            AND cid_res.cod_cidades = c.COD_CID_RES 
            AND cid_res.uf = c.UF_RES 
        LEFT JOIN cidades cid_com ON 1=1
            AND cid_com.cod_cidades = c.COD_CID_COM 
            AND cid_com.uf = c.UF_COM 
        LEFT JOIN cidades cid_cob ON 1=1
            AND cid_cob.cod_cidades = c.COD_CID_COBRANCA  
            AND cid_cob.uf = c.UF_COBRANCA
        LEFT JOIN CORES_EXTERNAS ce_veic ON 1=1          
            AND ce_veic.COR_EXTERNA = v.COR_EXTERNA
        LEFT JOIN CORES_EXTERNAS ce_ped ON 1=1         
            AND ce_ped.COR_EXTERNA = vped.COR_EXTERNA
        LEFT JOIN CORES_EXTERNAS ce_fic ON 1=1         
            AND ce_fic.COR_EXTERNA = pfd.COR_EXTERNA
        LEFT JOIN crm_eventos cev ON 1=1            
            AND cev.COD_PROPOSTA = vp.COD_PROPOSTA
            AND cev.status <> 'D'
            AND cev.cod_tipo_evento IN (829, 831,795,793,797,799,819,821,825,785,807,827,835,815,817,823,810,812)
        LEFT JOIN crm_andamento ca ON ca.COD_ANDAMENTO = cev.COD_ANDAMENTO
        WHERE 1=1
            AND vp.status_proposta = 'A'
            AND TRUNC(vp.EMISSAO) BETWEEN TO_DATE('{initial_date}', 'YYYY-MM-DD') AND TO_DATE('{final_date}', 'YYYY-MM-DD')
        ORDER BY pm.DESCRICAO_MODELO
    """
    cur.execute(query)
    result = cur.fetchall()
    columns = [desc[0] for desc in cur.description]
    df_abertos = pd.DataFrame(result, columns=columns)
    df_abertos['DATA_PROPOSTA'] = pd.to_datetime(df_abertos['DATA_PROPOSTA'], errors='coerce').dt.strftime('%d/%m/%Y').fillna('-')
    df_abertos['EVENTO'] = df_abertos['EVENTO'].fillna('')
    df_abertos['LINK_EVENTO'] = df_abertos.apply(
        lambda row: f"https://app.caiuas.com.br/crm/eventos/{row['EVENTO']}" if row['EVENTO'] else '', axis=1
    )
    
    query = f"""
        SELECT 
            to_char(v.COD_PROPOSTA) COD_PROPOSTA,
            vp.EMISSAO data_proposta, 
            eu.NOME_COMPLETO nome_vendedor, 
            pm.DESCRICAO_MODELO modelo, 
            ce.DESCRICAO cor, 
            v.CHASSI_COMPLETO, 
            c.NOME nome_cliente,
            CASE 
                WHEN c.NUMERO_MSG_TXT_INST IS NOT NULL THEN c.PREFIXO_MSG_TXT_INST || c.NUMERO_MSG_TXT_INST
                WHEN c.TELEFONE_CEL IS NOT NULL THEN c.PREFIXO_CEL || c.TELEFONE_CEL
                WHEN c.TELEFONE_res IS NOT NULL THEN c.PREFIXO_RES || c.TELEFONE_res
                WHEN c.TELEFONE_COM IS NOT NULL THEN c.PREFIXO_COM || c.TELEFONE_COM
            END AS telefone,
            c.EMAIL_NFE email,
            CASE 
                WHEN v.novo_usado = 'U' THEN 'Usado'
                ELSE
                    'Novo'
            END novo_usado,
            CASE
                WHEN c.COD_CID_COM <> NULL THEN cid_com.DESCRICAO
                WHEN c.COD_CID_RES <> NULL THEN cid_res.DESCRICAO 
                ELSE cid_cob.DESCRICAO 
            END cidade,
            'Faturado' Status,
            concat(cev2.cod_empresa, cev2.COD_EVENTO) evento,
            ca2.ANDAMENTO andamento,
            CASE 
                WHEN cev2.COD_PROPOSTA IS NOT NULL THEN 'Super quente'
                WHEN cev2.status = 'D' THEN 'Frio'
                WHEN cev2.status = 'E' AND cev2.COD_MOTIVO_PERDA IS NOT NULL THEN 'Frio'
                WHEN cev2.TERMOMETRO = 1 THEN 'Frio'
                WHEN cev2.TERMOMETRO = 2 THEN 'Morno'
                WHEN cev2.TERMOMETRO = 3 THEN 'Quente'
                ELSE 'Não classificado'
            END AS termometro
        FROM veiculos v 
        LEFT JOIN produtos pr ON 1=1
            AND pr.COD_PRODUTO = v.COD_PRODUTO 
        LEFT JOIN CORES_EXTERNAS ce ON 1=1
            AND ce.COR_EXTERNA = v.COR_EXTERNA 
        LEFT JOIN produtos_modelos pm ON 1=1
            AND pm.COD_PRODUTO = v.COD_PRODUTO 
            AND pm.COD_MODELO = v.COD_MODELO 
        LEFT JOIN VEICULOS_PROPOSTAS vp ON 1=1
            --AND vp.COD_PROPOSTA = v.COD_PROPOSTA OR vp.COD_PROPOSTA = v.COD_PROPOSTA_INTERNET 
            AND vp.CHASSI_RESUMIDO = v.CHASSI_RESUMIDO 
            AND vp.STATUS_PROPOSTA <> 'C'
        LEFT JOIN clientes c ON c.COD_CLIENTE = vp.COD_CLIENTE 
        LEFT JOIN patio p ON 1=1
            AND p.COD_PATIO = v.COD_PATIO
        LEFT JOIN EMPRESAS_USUARIOS eu ON 1=1
            AND eu.NOME = vp.VENDEDOR 
        LEFT JOIN empresas_usuarios eu2 ON 1=1
            AND eu2.nome = vp.QUEM_APROVOU 
        LEFT JOIN empresas e ON 1=1
            AND e.cod_empresa = v.COD_EMPRESA 
        LEFT JOIN cidades cid_res ON 1=1
            AND cid_res.cod_cidades = c.COD_CID_RES 
            AND cid_res.uf = c.UF_RES 
        LEFT JOIN cidades cid_com ON 1=1
            AND cid_com.cod_cidades = c.COD_CID_COM 
            AND cid_com.uf = c.UF_COM 
        LEFT JOIN cidades cid_cob ON 1=1
            AND cid_cob.cod_cidades = c.COD_CID_COBRANCA  
            AND cid_cob.uf = c.UF_COBRANCA
        LEFT JOIN crm_eventos cev2 ON 1=1
            AND cev2.COD_PROPOSTA = v.COD_PROPOSTA
            AND cev2.status <> 'D'
            AND cev2.cod_tipo_evento IN (829, 831,795,793,797,799,819,821,825,785,807,827,835,815,817,823,810,812)
        LEFT JOIN crm_andamento ca2 ON ca2.COD_ANDAMENTO = cev2.COD_ANDAMENTO
        WHERE v.status = 'V'
            AND vp.status_proposta = 'V'
            --AND v.cod_proposta <> 0
            --AND v.cod_proposta IS NOT NULL
            AND TRUNC(vp.EMISSAO) BETWEEN TO_DATE('{initial_date}', 'YYYY-MM-DD') AND TO_DATE('{final_date}', 'YYYY-MM-DD')
        ORDER BY pm.DESCRICAO_MODELO
    """
    cur.execute(query)
    result = cur.fetchall()
    columns = [desc[0] for desc in cur.description]
    df_faturados = pd.DataFrame(result, columns=columns)
    df_faturados['DATA_PROPOSTA'] = pd.to_datetime(df_faturados['DATA_PROPOSTA'], errors='coerce').dt.strftime('%d/%m/%Y').fillna('-')
    df_faturados['EVENTO'] = df_faturados['EVENTO'].fillna('')
    df_faturados['LINK_EVENTO'] = df_faturados.apply(
        lambda row: f"https://app.caiuas.com.br/crm/eventos/{row['EVENTO']}" if row['EVENTO'] else '', axis=1
    )
    
    
    ndf = pd.concat([df_abertos, df_faturados], ignore_index=True)
    ndf = ndf.sort_values(by='DATA_PROPOSTA', key=lambda x: pd.to_datetime(x, format='%d/%m/%Y', errors='coerce'), ascending=True)
    # replace none to ''
    ndf = ndf.fillna('')
    
    # fechando conexão
    cur.close()
    conn.close()
    
    ndf_download = ndf.rename(columns={
        'COD_PROPOSTA': 'Proposta',
        'DATA_PROPOSTA': 'Data Proposta',
        'NOME_VENDEDOR': 'Vendedor',
        'MODELO': 'Modelo',
        'COR': 'Cor',
        'CHASSI_COMPLETO': 'Chassi',
        'NOME_CLIENTE': 'Cliente',
        'NOVO_USADO': 'Novo/Usado',
        'CIDADE': 'Cidade',
        'STATUS': 'Status',
        'ANDAMENTO': 'Andamento',
        'TERMOMETRO': 'Temperatura',
        'LINK_EVENTO': 'Evento NBS'
    })
    import io
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    
    _xlsx_buffer = io.BytesIO()
    ndf_download.to_excel(_xlsx_buffer, index=False, engine='openpyxl')
    _xlsx_buffer.seek(0)
    
    wb = load_workbook(_xlsx_buffer)
    ws = wb.active
    
    # Formatar coluna "Data Proposta" como data real no Excel
    from datetime import datetime as _dt
    date_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == 'Data Proposta':
            date_col_idx = idx
            break
    if date_col_idx:
        for row in ws.iter_rows(min_row=2, min_col=date_col_idx, max_col=date_col_idx):
            for cell in row:
                if cell.value and cell.value != '-':
                    try:
                        cell.value = _dt.strptime(str(cell.value), '%d/%m/%Y')
                        cell.number_format = 'DD/MM/YYYY'
                    except ValueError:
                        pass
    
    # Criar tabela nomeada "Propostas"
    max_row = ws.max_row
    max_col = ws.max_column
    table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName="Propostas", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                          showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    
    # Redimensionar largura das colunas automaticamente
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)
    
    _xlsx_buffer = io.BytesIO()
    wb.save(_xlsx_buffer)
    _xlsx_buffer.seek(0)
    
    st.download_button(
        label="📥 Baixar planilha",
        data=_xlsx_buffer,
        file_name="propostas.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    
    st.dataframe(
        ndf_download,
        hide_index=True,
        use_container_width=True,
        height=800,
        column_config={
            "Andamento": st.column_config.TextColumn("Andamento"),
            "Temperatura": st.column_config.TextColumn("Temperatura"),
            "Evento NBS": st.column_config.LinkColumn("Evento NBS", display_text="Abrir"),
        }
    )
    
    
    
    